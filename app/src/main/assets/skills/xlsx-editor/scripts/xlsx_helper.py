"""Xlsx-editor helper - programmatic ops on .xlsx files.

Subcommands:
  inspect <file.xlsx> [--rows N] [--sheet NAME]   - per-sheet header + preview;
                                                    shows BOTH formula string
                                                    and cached value when
                                                    available.
  extract <file.xlsx> [--sheet NAME]              - dump cell values as CSV-
                                                    like text (cached values
                                                    when present, formulas
                                                    otherwise).
  set-cell <file> --sheet S --cell A1             - set a single cell. Pick ONE
        ( --text T | --number N | --formula F )     of --text / --number /
        [--numfmt FMT] [--out OUT]                  --formula. Cell font/fill/
                                                    alignment/border are
                                                    PRESERVED. --numfmt sets
                                                    the number format string
                                                    (e.g. '#,##0', '0.0%',
                                                    '$#,##0.00').
  set-range <file> SHEET A1 "v1,v2,.." [..]       - bulk write a rectangular
                                                    area; one CSV string per
                                                    ROW; '=' prefix = formula;
                                                    numbers auto-typed.
  new-from-outline <outline.json> --out OUT       - build a multi-sheet xlsx
        [--auto-color]                              from outline; with
                                                    --auto-color applies the
                                                    financial color standard
                                                    (blue=hardcoded,
                                                    black=formula).

Design:
  * Formula-first: any string starting with '=' is treated as a formula
    (preserves Excel recalculation behavior).
  * Style preservation on edit: set-cell only touches .value (and optionally
    .number_format), never destroys font / fill / alignment / border.
  * NEVER round-trip through openpyxl with `data_only=True` for editing -- it
    would replace formulas with cached values. Default load preserves them.

Dependency: openpyxl only.
"""
import sys
import os
import json
import argparse
from copy import copy, deepcopy

try:
    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# ===========================================================================
# Financial color standard
# ===========================================================================
COLOR_HARDCODED = "0000FF"   # blue for hardcoded inputs
COLOR_FORMULA   = "000000"   # black for formulas
COLOR_XSHEET    = "00B050"   # green for cross-sheet references


# ===========================================================================
# Helpers
# ===========================================================================
def _is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def _coerce(v):
    """Convert a raw JSON cell value to (cell_value, is_formula)."""
    if isinstance(v, dict):
        if "formula" in v:
            return ("=" + v["formula"].lstrip("="), True)
        if "value" in v:
            return (v["value"], _is_formula(v["value"]))
    if _is_formula(v):
        return (v, True)
    return (v, False)


def _preview(v, max_len=40):
    s = "" if v is None else str(v)
    s = s.replace("\n", "\\n").replace("\r", "")
    if len(s) > max_len:
        s = s[:max_len - 1] + "…"
    return s


# ===========================================================================
# Subcommand: inspect
# ===========================================================================
def cmd_inspect(args):
    wb_f = load_workbook(args.file, data_only=False)
    try:
        wb_v = load_workbook(args.file, data_only=True)
    except Exception:
        wb_v = None

    sheets = [args.sheet] if args.sheet else wb_f.sheetnames
    print(f"[INSPECT] {args.file}")
    print(f"  sheets: {wb_f.sheetnames}")
    print()
    for name in sheets:
        if name not in wb_f.sheetnames:
            print(f"  (sheet '{name}' not found)")
            continue
        ws = wb_f[name]
        ws_v = wb_v[name] if wb_v is not None and name in wb_v.sheetnames else None
        print(f"─── Sheet '{name}' ───")
        print(f"  dimensions: {ws.dimensions}   "
              f"max_row={ws.max_row}   max_col={ws.max_column}")
        # Show column widths
        widths = []
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            cd = ws.column_dimensions.get(letter)
            if cd and cd.width:
                widths.append(f"{letter}={cd.width:.1f}")
        if widths:
            print(f"  widths: {' '.join(widths)}")
        # Print up to args.rows rows
        n = min(args.rows, ws.max_row)
        print(f"  preview (first {n} rows):")
        for r in range(1, n + 1):
            row_repr = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                addr = cell.coordinate
                v = cell.value
                if _is_formula(v):
                    cached = ws_v.cell(r, c).value if ws_v is not None else None
                    if cached is not None:
                        row_repr.append(f"{addr}={v!r}→{cached!r}")
                    else:
                        row_repr.append(f"{addr}={v!r}")
                elif v is not None:
                    nf = cell.number_format
                    if nf and nf != "General":
                        row_repr.append(f"{addr}={_preview(v)} [{nf}]")
                    else:
                        row_repr.append(f"{addr}={_preview(v)}")
            if row_repr:
                print("    " + "  ".join(row_repr))
        # Merged cells
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f"  merged: {[str(m) for m in merged]}")
        print()
    _next_step([
        "You now have sheet structure (dimensions, headers, formulas + cached).",
        f"Edit a cell:    python xlsx_helper.py set-cell {args.file} --sheet S --cell A1 --formula \"=...\"",
        f"Bulk write 2D:  python xlsx_helper.py set-range {args.file} S A1 \"v1,v2,...\" \"v3,v4,...\"",
    ])


# ===========================================================================
# Subcommand: extract
# ===========================================================================
def cmd_extract(args):
    wb = load_workbook(args.file, data_only=True)
    sheets = [args.sheet] if args.sheet else wb.sheetnames
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        print(f"### Sheet: {name}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            # Trim trailing empties
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                print("\t".join(cells))
        print()
    _next_step([
        "Tab-separated dump above (cached values when available).",
        f"To see formulas + addresses: python xlsx_helper.py inspect {args.file}",
    ])


# ===========================================================================
# Subcommand: set-cell
# ===========================================================================
def cmd_set_cell(args):
    """Set a single cell. Simplified signature:

        set-cell FILE SHEET A1 VALUE [--numfmt FMT]

    VALUE intent is auto-detected:
      * '=' prefix            -> formula
      * int/float literal     -> number
      * anything else         -> text

    Legacy explicit flags (--text / --number / --formula) still work as
    overrides for the rare case auto-detection gets in the way.
    """
    # Compat shim: positional sheet_pos/cell_pos fill legacy attrs.
    if not args.sheet:
        args.sheet = getattr(args, "sheet_pos", None)
    if not args.cell:
        args.cell = getattr(args, "cell_pos", None)
    if not args.sheet or not args.cell:
        raise ValueError(
            "set-cell: SHEET and CELL required. Usage:\n"
            "  set-cell FILE SHEET A1 VALUE [--numfmt FMT]"
        )
    # Resolve the primary VALUE. Priority: explicit flag > positional.
    if args.formula is not None:
        v = args.formula if args.formula.startswith("=") else "=" + args.formula
    elif args.number is not None:
        try:
            v = int(args.number)
            if str(v) != str(args.number):
                v = float(args.number)
        except ValueError:
            v = float(args.number)
    elif args.text is not None:
        v = args.text
    else:
        raw = getattr(args, "value_pos", None)
        if raw is None:
            raise ValueError(
                "set-cell: VALUE required. Usage:\n"
                "  set-cell FILE SHEET A1 VALUE             "
                "(auto-detect: '=' -> formula, number -> number, else text)\n"
                "Or legacy explicit form:\n"
                "  set-cell FILE SHEET A1 (--text T | --number N | --formula F)"
            )
        # Auto-detect: formula > number > text.
        s = str(raw).strip()
        if s.startswith("="):
            v = s
        else:
            try:
                v = int(s)
                if str(v) != s:  # leading zeros etc -> keep as float/text
                    raise ValueError
            except ValueError:
                try:
                    v = float(s)
                except ValueError:
                    v = raw  # keep original (preserves whitespace if any)
    wb = load_workbook(args.file, data_only=False)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"sheet '{args.sheet}' not found in {wb.sheetnames}")
    ws = wb[args.sheet]
    cell = ws[args.cell]   # accepts e.g. 'A1', 'B12'
    # Setting .value alone preserves font/fill/alignment/border/number_format.
    cell.value = v
    if args.numfmt:
        cell.number_format = args.numfmt
    out = args.out or args.file
    wb.save(out)
    print(f"[OK] set-cell {args.sheet}!{args.cell} = {v!r} -> {out}")
    _next_step([
        f"Cell {args.sheet}!{args.cell} updated.",
        f"Verify:  python xlsx_helper.py inspect {out} --sheet {args.sheet}",
        "Continue with more set-cell or set-range calls.",
    ])


# ===========================================================================
# Subcommand: set-range
# ===========================================================================
def cmd_set_range(args):
    """Bulk-write a rectangular area using CSV-string ROWs (no JSON file).

    Canonical:  set-range FILE SHEET A1 "v1,v2,v3" ["v4,v5,v6" ...]
    Each positional ROW is a CSV string: '=' prefix => formula, numeric-like
    strings get auto-coerced to int/float, empty cells are skipped (preserves
    existing cell values), everything else becomes text. Interface mirrors
    add-sheet/add-row so the model can reuse the same mental model.
    """
    if not args.sheet:
        args.sheet = getattr(args, "sheet_pos", None)
    if not args.start:
        args.start = getattr(args, "start_pos", None)
    rows_raw = list(getattr(args, "rows_pos", None) or [])

    if not args.sheet or not args.start or not rows_raw:
        raise ValueError(
            "set-range: SHEET, ANCHOR, and at least one ROW required. Usage:\n"
            "  set-range FILE SHEET A1 \"v1,v2,v3\" [\"v4,v5,v6\" ...]\n"
            "  '=' prefix = formula; numeric strings auto-typed; "
            "empty cell keeps existing value."
        )

    # Guard against callers still thinking in JSON. The prior contract used a
    # JSON file path here; bounce them to the new CSV form with an actionable
    # message instead of silently writing literal "[[...]]" into a cell.
    first = rows_raw[0].strip()
    if first.startswith(("[", "{")) or first.lower().endswith(".json"):
        raise ValueError(
            "set-range: rows are CSV strings now, not JSON.\n"
            "  New usage:\n"
            "    set-range FILE SHEET A1 \"v1,v2,v3\" \"v4,v5,v6\"\n"
            "  '=' prefix = formula; numbers auto-typed; "
            "comma is separator."
        )

    wb = load_workbook(args.file, data_only=False)
    if args.sheet not in wb.sheetnames:
        raise ValueError(
            f"set-range: sheet '{args.sheet}' not found. "
            f"Existing sheets: {wb.sheetnames}"
        )
    ws = wb[args.sheet]
    try:
        start_cell = ws[args.start]
    except Exception:
        raise ValueError(
            f"set-range: ANCHOR '{args.start}' is not a valid cell address "
            f"(use e.g. 'A1', 'B2', 'D10')."
        )
    r0, c0 = start_cell.row, start_cell.column

    n_cells = 0
    max_cols = 0
    for ri, row_str in enumerate(rows_raw):
        cells = _split_csv(row_str)
        if len(cells) > max_cols:
            max_cols = len(cells)
        for ci, raw in enumerate(cells):
            # Empty string => skip (don't clobber existing value).
            if raw == "":
                continue
            cell = ws.cell(r0 + ri, c0 + ci)
            cell.value = _coerce_scalar(raw)
            n_cells += 1
    out = args.out or args.file
    wb.save(out)
    print(f"[OK] set-range {args.sheet}!{args.start} wrote {n_cells} cells "
          f"over {len(rows_raw)} row(s) x up to {max_cols} col(s) -> {out}")


# ===========================================================================
# Subcommand: new-from-outline
# ===========================================================================
def _apply_auto_color(cell):
    """Apply blue=hardcoded, black=formula coloring based on cell value."""
    v = cell.value
    if v is None:
        return
    color = COLOR_FORMULA if _is_formula(v) else COLOR_HARDCODED
    f = cell.font
    cell.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                     color=color)


def _validate_outline(outline):
    issues = []
    if not isinstance(outline, dict):
        return ["outline must be a JSON object"]
    sheets = outline.get("sheets", [])
    if not isinstance(sheets, list):
        return ["'sheets' must be an array"]
    if not sheets:
        return ["'sheets' is empty"]
    seen_names = set()
    for i, sd in enumerate(sheets):
        prefix = f"sheets[{i}]"
        if not isinstance(sd, dict):
            issues.append(f"{prefix}: must be an object"); continue
        name = sd.get("name", "")
        if not name:
            issues.append(f"{prefix}: 'name' is required")
        elif name in seen_names:
            issues.append(f"{prefix}: duplicate sheet name '{name}'")
        else:
            seen_names.add(name)
        if len(name) > 31:
            issues.append(f"{prefix}: sheet name '{name}' >31 chars (Excel limit)")

        headers = sd.get("headers", [])
        rows = sd.get("rows", [])
        totals = sd.get("totals_row")
        if headers and not isinstance(headers, list):
            issues.append(f"{prefix}: 'headers' must be an array")
        if rows and not isinstance(rows, list):
            issues.append(f"{prefix}: 'rows' must be a 2D array")
        # Width check vs headers
        nh = len(headers)
        for ri, row in enumerate(rows):
            if isinstance(row, list) and nh and len(row) > nh:
                issues.append(f"{prefix}: rows[{ri}] has {len(row)} cols, "
                              f"headers has {nh}")
        if totals and isinstance(totals, list) and nh and len(totals) > nh:
            issues.append(f"{prefix}: totals_row has {len(totals)} cols, "
                          f"headers has {nh}")
        # number_formats keys must be column letters
        for col_letter in (sd.get("number_formats") or {}).keys():
            if not col_letter.isalpha() or not col_letter.isupper():
                issues.append(f"{prefix}: number_formats key '{col_letter}' "
                              f"must be uppercase column letter (A, B, AA…)")
        # Detect "Python-precomputed totals" anti-pattern (totals row with all
        # hard numbers, no formulas) — financial-model rule violation.
        if totals and isinstance(totals, list):
            data_cols = sum(1 for v in totals[1:]
                            if isinstance(v, (int, float)))
            formula_cols = sum(1 for v in totals[1:]
                               if isinstance(v, str) and v.startswith("="))
            if data_cols >= 2 and formula_cols == 0:
                issues.append(f"{prefix}: totals_row uses hardcoded numbers "
                              f"only; use Excel formulas (e.g. '=SUM(B2:B5)')")
    return issues


# ===========================================================================
# Built-in example outlines.
# ===========================================================================
EXAMPLES_XLSX = {
    "sales": {
        "_desc": "Regional sales data with totals row + assumptions sheet.",
        "outline": {
            "sheets": [
                {
                    "name": "Sales",
                    "title": "Q4 2024 Regional Sales",
                    "headers": ["Region", "Q1", "Q2", "Q3", "Q4", "Total"],
                    "rows": [
                        ["North America", 1200000, 1350000, 1500000, 1750000, "=SUM(B4:E4)"],
                        ["Europe",        800000,  900000,  950000,  1100000, "=SUM(B5:E5)"],
                        ["APAC",          400000,  500000,  600000,  800000,  "=SUM(B6:E6)"],
                        ["LATAM",         200000,  220000,  250000,  280000,  "=SUM(B7:E7)"]
                    ],
                    "totals_row": ["Total", "=SUM(B4:B7)", "=SUM(C4:C7)",
                                   "=SUM(D4:D7)", "=SUM(E4:E7)", "=SUM(F4:F7)"],
                    "column_widths": {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 16},
                    "number_formats": {"B": "$#,##0", "C": "$#,##0", "D": "$#,##0",
                                       "E": "$#,##0", "F": "$#,##0"},
                    "freeze_top_row": True
                },
                {
                    "name": "Assumptions",
                    "title": "Key Assumptions",
                    "headers": ["Item", "Value", "Source"],
                    "rows": [
                        ["Growth Rate",   0.18, "Q3 Board Memo"],
                        ["Op Margin",     0.15, "FY2025 Plan"],
                        ["FX EUR/USD",    1.07, "Bloomberg 2024-12-31"]
                    ],
                    "column_widths": {"A": 22, "B": 12, "C": 28}
                }
            ]
        }
    },

    "financial_model": {
        "_desc": "3-statement-style mini-model: assumptions sheet drives revenue projection.",
        "outline": {
            "sheets": [
                {
                    "name": "Assumptions",
                    "title": "Model Assumptions",
                    "headers": ["Item", "Value", "Source"],
                    "rows": [
                        ["Base Revenue ($M)",     5.2, "Q4 2024 actual"],
                        ["Growth Rate (annual)",  0.18, "Board approved"],
                        ["Gross Margin",          0.62, "Q4 actual"],
                        ["Op Expense ($M)",       2.4, "FY2025 plan"]
                    ],
                    "column_widths": {"A": 24, "B": 12, "C": 28},
                    "number_formats": {"B": "#,##0.00"}
                },
                {
                    "name": "Projection",
                    "title": "5-Year Revenue Projection",
                    "headers": ["Metric", "FY2025", "FY2026", "FY2027", "FY2028", "FY2029"],
                    "rows": [
                        ["Revenue ($M)",
                         "=Assumptions!B4*(1+Assumptions!B5)",
                         "=B4*(1+Assumptions!$B$5)",
                         "=C4*(1+Assumptions!$B$5)",
                         "=D4*(1+Assumptions!$B$5)",
                         "=E4*(1+Assumptions!$B$5)"],
                        ["Gross Profit",
                         "=B4*Assumptions!$B$6", "=C4*Assumptions!$B$6",
                         "=D4*Assumptions!$B$6", "=E4*Assumptions!$B$6",
                         "=F4*Assumptions!$B$6"],
                        ["Op Income",
                         "=B5-Assumptions!$B$7", "=C5-Assumptions!$B$7",
                         "=D5-Assumptions!$B$7", "=E5-Assumptions!$B$7",
                         "=F5-Assumptions!$B$7"],
                        ["Op Margin",
                         "=B6/B4", "=C6/C4", "=D6/D4", "=E6/E4", "=F6/F4"]
                    ],
                    "column_widths": {"A": 16, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14},
                    "number_formats": {"B": "$#,##0.0", "C": "$#,##0.0",
                                       "D": "$#,##0.0", "E": "$#,##0.0",
                                       "F": "$#,##0.0"}
                }
            ]
        }
    },

    "budget": {
        "_desc": "Department budget tracker (planned vs actual, variance formula).",
        "outline": {
            "sheets": [
                {
                    "name": "Budget",
                    "title": "FY2025 Engineering Budget",
                    "headers": ["Category", "Planned", "Actual", "Variance", "Variance %"],
                    "rows": [
                        ["Salaries",       2400000, 2380000, "=B4-C4", "=D4/B4"],
                        ["Cloud + SaaS",   480000,  520000,  "=B5-C5", "=D5/B5"],
                        ["Hardware",       180000,  195000,  "=B6-C6", "=D6/B6"],
                        ["Travel",         60000,   45000,   "=B7-C7", "=D7/B7"],
                        ["Training",       40000,   38000,   "=B8-C8", "=D8/B8"]
                    ],
                    "totals_row": ["Total", "=SUM(B4:B8)", "=SUM(C4:C8)",
                                   "=SUM(D4:D8)", "=D9/B9"],
                    "column_widths": {"A": 18, "B": 14, "C": 14, "D": 14, "E": 12},
                    "number_formats": {"B": "$#,##0", "C": "$#,##0",
                                       "D": "$#,##0;($#,##0);-", "E": "0.0%"}
                }
            ]
        }
    }
}


# ===========================================================================
# _next_step: intentional no-op.
#
# The helper has NO way to know the task's overall goal (is the model going
# to draw more sheets? pivot to docx? go back to the web for more data?).
# Any "continue with another add-*" / "finish: inspect" suggestion is just
# a guess that pollutes the model's attention.
#
# Scripts should only:
#   * report WHAT HAPPENED    -> [OK] + [STATE]
#   * report errors           -> [ERROR] + targeted recovery hint
#   * provide param reference -> [HINT for cmd] (just-run command's params)
# Direction belongs to the model / user prompt, not the helper.
# Kept as a no-op so call sites don't have to be refactored.
# ===========================================================================
def _next_step(lines):  # noqa: ARG001 - intentionally unused
    return


def cmd_examples(args):
    # [DEPRECATED] outline-template path. Prefer the incremental block API:
    #   python xlsx_helper.py catalog
    #   python xlsx_helper.py new-book FILE.xlsx [--auto-color]
    #   python xlsx_helper.py add-sheet / add-assumptions / add-dashboard ...
    print("[DEPRECATED] `examples` is the legacy outline-template path. "
          "Prefer `catalog` + `new-book` + `add-*` for incremental, "
          "per-sheet construction.")
    name = args.name
    if not name or name == "list":
        print("Available outline examples (use: examples <name> [--out FILE]):")
        for k, v in EXAMPLES_XLSX.items():
            print(f"  {k:<18}  {v['_desc']}")
        _next_step([
            "Pick a template and dump its outline JSON to a file:",
            "  python xlsx_helper.py examples <name> --out ${WORKSPACE}/outline.json",
            "Or take the fast path (no outline editing):",
            "  python xlsx_helper.py create --template <name> --out ${WORKSPACE}/book.xlsx",
        ])
        return
    if name not in EXAMPLES_XLSX:
        raise ValueError(f"unknown example '{name}'. choices: "
                         f"{list(EXAMPLES_XLSX)} or 'list'")
    outline_json = json.dumps(EXAMPLES_XLSX[name]["outline"],
                              ensure_ascii=False, indent=2)
    out_path = getattr(args, "out", None)
    if out_path:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(outline_json)
        print(f"[OK] example '{name}' outline written to {out_path}")
        print(f"[WARN] This file is the UNMODIFIED '{name}' template placeholder. "
              f"If you pass it straight to new-from-outline, the build will be REJECTED.")
        _next_step([
            f"REQUIRED next: overwrite {out_path} with your real task data using write_file.",
            "  - Edit each sheets[i]: 'name', 'title', 'headers', 'rows', 'totals_row'.",
            "    Keep 'number_formats' / 'column_widths' semantics sane; formulas start with '='.",
            f"Then validate: python xlsx_helper.py new-from-outline {out_path} --dry-run",
            f"Finally build: python xlsx_helper.py new-from-outline {out_path} --out ${{WORKSPACE}}/book.xlsx",
        ])
    else:
        print(outline_json)
        _next_step([
            "Save the JSON above to a file, or re-run with --out FILE to write directly.",
            "Then validate/build with new-from-outline.",
        ])


def _load_json(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def _outline_signature(outline):
    """Stable content hash of an outline, used to detect unmodified templates."""
    return json.dumps(outline, sort_keys=True, ensure_ascii=False)


_EXAMPLE_SIGNATURES = {
    name: _outline_signature(v["outline"]) for name, v in EXAMPLES_XLSX.items()
}


def _detect_unmodified_template(outline):
    """Return example name if outline matches a built-in example byte-for-byte.

    Used to reject `new-from-outline` calls where the LLM forgot to edit the
    example outline produced by `examples NAME --out`, which would otherwise
    build a workbook containing the template's placeholder copy instead of
    the real task data.
    """
    sig = _outline_signature(outline)
    for name, ref_sig in _EXAMPLE_SIGNATURES.items():
        if sig == ref_sig:
            return name
    return None


def _build_from_outline(outline, out_path, auto_color=False):
    """Core xlsx builder; shared by new-from-outline and create.

    Returns (sheet_count, auto_color_applied).
    """
    sheets_def = outline.get("sheets", [])
    if not sheets_def:
        raise ValueError("outline must include non-empty 'sheets' array")

    wb = Workbook()
    # Remove default sheet; re-create explicitly in user order.
    wb.remove(wb.active)

    for sd in sheets_def:
        name = sd.get("name", f"Sheet{len(wb.sheetnames) + 1}")
        ws = wb.create_sheet(title=name)

        headers = sd.get("headers")
        rows = sd.get("rows", [])
        totals = sd.get("totals_row")
        widths = sd.get("column_widths", {})
        numfmts = sd.get("number_formats", {})
        header_fill = sd.get("header_fill", "1F3864")  # dark blue default
        freeze = sd.get("freeze_top_row", True)

        r = 1
        # Title block
        if sd.get("title"):
            ws.cell(r, 1).value = sd["title"]
            ws.cell(r, 1).font = Font(bold=True, size=14)
            r += 1
            r += 1  # blank line

        header_row = None
        if headers:
            header_row = r
            for ci, h in enumerate(headers, start=1):
                cell = ws.cell(r, ci)
                cell.value = h
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=header_fill)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            r += 1

        # Data rows
        for row_data in rows:
            for ci, raw in enumerate(row_data, start=1):
                if raw is None:
                    continue
                cell = ws.cell(r, ci)
                v, _ = _coerce(raw)
                cell.value = v
            r += 1

        # Totals row
        if totals:
            for ci, raw in enumerate(totals, start=1):
                if raw is None:
                    continue
                cell = ws.cell(r, ci)
                v, _ = _coerce(raw)
                cell.value = v
                cell.font = Font(bold=True)
                # Top border for accounting style
                cell.border = Border(top=Side(style="thin"))
            r += 1

        # Column widths
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = float(w)

        # Number formats
        for col_letter, fmt in numfmts.items():
            col_idx = column_index_from_string(col_letter)
            for rr in range(1, ws.max_row + 1):
                c = ws.cell(rr, col_idx)
                if c.value is not None and not isinstance(c.value, str):
                    c.number_format = fmt
                elif _is_formula(c.value):
                    c.number_format = fmt

        if freeze and header_row is not None:
            ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate

        # Auto color (financial standard)
        if auto_color:
            for row_cells in ws.iter_rows(min_row=(header_row + 1) if header_row else 1,
                                          values_only=False):
                for cell in row_cells:
                    if cell.value is None:
                        continue
                    _apply_auto_color(cell)

    wb.save(out_path)
    return len(wb.sheetnames), bool(auto_color)


def cmd_new_from_outline(args):
    # [DEPRECATED] outline-driven path; kept for backward compatibility.
    print("[DEPRECATED] `new-from-outline` builds a whole workbook in one "
          "shot from a JSON outline. Prefer the incremental API: "
          "`new-book` + repeated `add-*`.")
    outline_path = args.outline_pos or args.outline_opt
    if not outline_path:
        raise ValueError("outline path is required "
                         "(pass as positional or --outline PATH)")
    outline = _load_json(outline_path)

    # Guard: refuse to build when the outline is byte-for-byte identical to a
    # built-in example (the LLM forgot to edit it). Otherwise the workbook
    # would contain the template's placeholder copy, not real task data.
    tmpl = _detect_unmodified_template(outline)
    if tmpl:
        print(f"[ERROR] outline.json is the UNMODIFIED '{tmpl}' example template.")
        print(f"        Building now would produce a generic '{tmpl}' workbook, NOT your task data.")
        print(f"        You MUST edit the following fields to match the task before re-running:")
        print(f"          - each sheets[i]: 'name', 'title', 'headers', 'rows', 'totals_row'")
        print(f"          - keep 'number_formats' / 'column_widths' style intact")
        print(f"        Use write_file to overwrite {outline_path} with your customized outline,")
        print(f"        then re-run `new-from-outline`.")
        _next_step([
            f"Overwrite {outline_path} with real task data,",
            f"then validate: python xlsx_helper.py new-from-outline {outline_path} --dry-run",
        ])
        sys.exit(1)

    issues = _validate_outline(outline)
    if issues:
        print("[VALIDATION] issues found:")
        for it in issues:
            print(f"  - {it}")
    else:
        print("[VALIDATION] OK")

    if args.dry_run:
        print("[DRY-RUN] no file written.")
        if issues:
            _next_step([
                "Fix the issues listed above in the outline JSON, then re-run:",
                f"  python xlsx_helper.py new-from-outline {outline_path} --dry-run",
            ])
            sys.exit(1)
        _next_step([
            "Outline passed validation. To actually build the file, run:",
            f"  python xlsx_helper.py new-from-outline {outline_path} --out ${{WORKSPACE}}/book.xlsx",
        ])
        return

    if not args.out:
        raise ValueError("--out PATH is required unless --dry-run is set")
    n_sheets, ac = _build_from_outline(outline, args.out, args.auto_color)
    print(f"[OK] new-from-outline -> {args.out} "
          f"(sheets={n_sheets}, auto_color={ac})")
    _next_step([
        f"Workbook generated at {args.out}",
        f"Inspect:  python xlsx_helper.py inspect {args.out}",
        f"Edit:     python xlsx_helper.py set-cell {args.out} --sheet S --cell A1 --formula \"=...\"",
    ])


def cmd_create(args):
    """Fast path: generate an xlsx directly from a built-in template."""
    # [DEPRECATED] template-driven one-shot path.
    print("[DEPRECATED] `create` builds a whole workbook from a built-in "
          "template. Prefer the incremental API: `catalog` + `new-book` + "
          "`add-sheet` / `add-assumptions` / `add-dashboard`.")
    if args.outline_file:
        outline = _load_json(args.outline_file)
        src = f"outline file {args.outline_file}"
    else:
        if args.template not in EXAMPLES_XLSX:
            raise ValueError(f"unknown template '{args.template}'. "
                             f"choices: {list(EXAMPLES_XLSX)}")
        outline = deepcopy(EXAMPLES_XLSX[args.template]["outline"])
        src = f"built-in template '{args.template}'"
    issues = _validate_outline(outline)
    if issues:
        print("[VALIDATION] issues:")
        for it in issues:
            print(f"  - {it}")
    else:
        print("[VALIDATION] OK")
    n_sheets, ac = _build_from_outline(outline, args.out, args.auto_color)
    print(f"[OK] create -> {args.out} (from {src}; "
          f"sheets={n_sheets}, auto_color={ac})")
    _next_step([
        f"Workbook generated at {args.out}",
        f"Inspect:  python xlsx_helper.py inspect {args.out}",
        f"Edit:     python xlsx_helper.py set-cell {args.out} --sheet S --cell A1 --formula \"=...\"",
        "Custom content path: examples <name> --out outline.json "
        "-> edit fields -> new-from-outline <file> --out <book.xlsx>",
    ])


# ===========================================================================
# === Catalog / new-book / add-* incremental API (mirrors pptx/docx helpers)
# Per-sheet additive flow: every add-* command appends ONE sheet (data sheet /
# assumptions sheet / KPI dashboard) and immediately persists the workbook.
# Inputs that exceed the hard caps below are auto-truncated with `[WARN]`
# (these commands NEVER fail on overlong content).
#
# Layout convention used by every add-* command:
#   * if --title given: row 1 = title, row 2 blank, row 3 = headers,
#                       row 4+ = data, then totals (if any)
#   * if no --title:    row 1 = headers, row 2+ = data, then totals
# Use the corresponding absolute row numbers when writing formulas.
# ===========================================================================

# Hard caps for the incremental sheet-builder API. Sized for "one screen of
# Excel data" so the model splits big tables across multiple add-sheet calls.
_LIM = {
    "sheet_name": 31,        # Excel hard upper bound on sheet names
    "title": 200,            # title row text
    "cell": 500,             # any single cell value
    "headers_max": 20,       # max columns per add-sheet
    "rows_max": 200,         # max body rows per add-sheet call
    "kv_max": 50,            # max --kv entries per add-assumptions
    "kpi_max": 12,           # max --kpi entries per add-dashboard
}

_META_MARK_XLSX = "xlsx-helper-book"


def _meta_default():
    """Default workbook-level metadata stored in core_properties."""
    return {"sheet_count": 0, "auto_color": False}


def _meta_load(wb):
    """Load workbook metadata stamped by this helper, with fallbacks."""
    p = wb.properties
    raw = p.description or ""
    try:
        m = json.loads(raw) if raw else None
        if isinstance(m, dict) and p.subject == _META_MARK_XLSX:
            out = _meta_default()
            out.update({k: m[k] for k in out.keys() if k in m})
            out["sheet_count"] = int(out["sheet_count"])
            out["auto_color"] = bool(out["auto_color"])
            return out
    except Exception:
        pass
    fb = _meta_default()
    if p.subject != _META_MARK_XLSX:
        # Existing workbook not produced by this helper -- still allow add-*
        # but seed the sheet count from the actual workbook so [OK] is honest.
        try:
            fb["sheet_count"] = len(wb.sheetnames)
        except Exception:
            pass
    return fb


def _meta_save(wb, meta):
    """Persist meta back into core_properties (subject + description)."""
    p = wb.properties
    p.subject = _META_MARK_XLSX
    p.description = json.dumps(meta, ensure_ascii=False)


# --- input clipping (no-fail) ----------------------------------------------
def _clip_str(s, max_chars, *, where=None):
    """Truncate string to max_chars; print [WARN] if trimmed."""
    if s is None:
        return ""
    s = str(s)
    if len(s) <= max_chars:
        return s
    if where:
        print(f"[WARN] {where}: {len(s)} chars > limit {max_chars}, truncated")
    return s[: max_chars - 1] + "\u2026"


def _clip_list(lst, max_n, *, where=None):
    """Drop tail items beyond max_n; print [WARN] if dropped."""
    if not lst:
        return []
    if len(lst) <= max_n:
        return list(lst)
    if where:
        print(f"[WARN] {where}: kept first {max_n} of {len(lst)} items")
    return list(lst[:max_n])


def _split_csv(text, sep=","):
    """Split a comma-separated arg into stripped non-empty tokens."""
    if not text:
        return []
    return [t.strip() for t in str(text).split(sep) if t.strip()]


def _open_book(path):
    """Load a workbook; if missing, instruct caller to run new-book first.

    The error message also lists every existing .xlsx under the same
    directory so the LLM can detect filename-drift bugs (observed in
    log3.txt: model used 3 different workbook names in a single task --
    报价数据.xlsx, flash_dram_price.xlsx, 报价.xlsx -- and never realized
    they were all "the same workbook in its head"). Surfacing the real
    candidates lets the model self-correct on the very next step.
    """
    if not os.path.exists(path):
        # List sibling .xlsx files to help recover from filename drift.
        siblings = []
        try:
            parent_dir = os.path.dirname(os.path.abspath(path)) or "."
            if os.path.isdir(parent_dir):
                siblings = sorted(
                    f for f in os.listdir(parent_dir)
                    if f.lower().endswith(".xlsx") and not f.startswith("~$")
                )
        except OSError:
            siblings = []
        msg = "Workbook not found: " + str(path) + "\n"
        if siblings:
            msg += (
                "Existing .xlsx in the same folder (did you mean one of these?):\n"
                + "\n".join("  * " + s for s in siblings) + "\n"
            )
        else:
            msg += "(No other .xlsx files in the same folder.)\n"
        msg += (
            "Fix options:\n"
            "  (A) If you meant one of the files above, retry with the "
            "correct name.\n"
            "  (B) If this is a brand-new workbook, run first:\n"
            "      python xlsx_helper.py new-book " + str(path) + " --auto-color\n"
            "RULE: One task = ONE workbook filename. Pick a name once and "
            "do NOT change it across steps."
        )
        raise FileNotFoundError(msg)
    return load_workbook(path, data_only=False)


def _coerce_scalar(raw):
    """Coerce a string-form cell value to int/float when it looks numeric.

    Strings that start with '=' stay as formulas. Empty strings stay empty.
    Anything else with non-numeric content stays a string.
    """
    if raw is None:
        return None
    if not isinstance(raw, str):
        return raw
    if _is_formula(raw):
        return raw
    s = raw.strip()
    if s == "":
        return ""
    # Try int / float coercion -- be conservative (no leading '+' tolerated).
    try:
        if "." in s or "e" in s.lower():
            return float(s)
        return int(s)
    except ValueError:
        return raw


def _parse_col_pair(spec):
    """Parse 'B:#,##0' or 'A:18' into (col_letter_upper, value_str). On
    malformed input return (None, None) so caller can [WARN] and skip."""
    if not spec or ":" not in str(spec):
        return None, None
    col, val = str(spec).split(":", 1)
    col = col.strip().upper()
    val = val.strip()
    if not col.isalpha() or not val:
        return None, None
    return col, val


# ---------------------------------------------------------------------------
# Smart defaults (C-plan): auto-inferred numfmt + auto-fit widths so 90% of
# add-sheet calls do NOT need --numfmt / --width flags. Explicit user flags
# still win; inference only fills columns the user did not specify.
# ---------------------------------------------------------------------------

_CURRENCY_PREFIXES = ("$", "\u00a5", "\u20ac", "\uffe5")  # $ ¥ € ￥

_CLS_TO_FMT = {
    "currency2": "$#,##0.00",
    "currency0": "$#,##0",
    "percent":   "0.0%",
    "int":       "#,##0",
    "float2":    "#,##0.00",
}


def _classify_column(samples):
    """Classify a column by inspecting its raw (pre-coerce) string samples.

    Returns one of the _CLS_TO_FMT keys, or None if mixed/text.
    Rules (first match wins):
      - ALL samples start with a currency prefix -> currency0 / currency2
      - ALL samples end with '%' -> percent
      - ALL samples parse as int/float (after comma strip) -> int / float2
      - Anything else -> None (leave as plain text, no numfmt)
    """
    if not samples:
        return None
    # Currency: every sample begins with $ / ¥ / € / ￥
    if all(any(s.startswith(p) for p in _CURRENCY_PREFIXES) for s in samples):
        has_decimal = any("." in s for s in samples)
        return "currency2" if has_decimal else "currency0"
    # Percent: every sample ends with %
    if all(s.endswith("%") for s in samples):
        return "percent"
    # Plain numeric: comma-stripped must parse as float
    has_decimal = False
    for s in samples:
        t = s.replace(",", "").lstrip("-+")
        if "." in t:
            has_decimal = True
            try:
                float(t)
            except ValueError:
                return None
        else:
            if not t.isdigit():
                return None
    return "float2" if has_decimal else "int"


def _coerce_with_classification(raw, col_cls):
    """Column-aware coercion that strips $/%/, so numfmt actually renders.

    When col_cls is None (text column) falls back to _coerce_scalar.
    For percent, values are divided by 100 since Excel stores 0.1 == '10%'.
    """
    if raw is None:
        return None
    if not isinstance(raw, str):
        return raw
    if _is_formula(raw):
        return raw
    s = raw.strip()
    if not s:
        return ""
    if not col_cls:
        return _coerce_scalar(raw)
    try:
        if col_cls.startswith("currency"):
            for prefix in _CURRENCY_PREFIXES:
                if s.startswith(prefix):
                    s = s[len(prefix):].strip()
                    break
            s = s.replace(",", "")
            return float(s) if "." in s else int(s)
        if col_cls == "percent":
            s = s.rstrip("%").strip().replace(",", "")
            return float(s) / 100.0  # Excel stores percent as fraction
        s = s.replace(",", "")
        if col_cls == "int":
            return int(s)
        return float(s) if "." in s else int(s)
    except ValueError:
        return raw  # bail out to raw string if parse breaks


def _auto_infer_numfmts(headers, rows, totals, user_numfmts):
    """Return (inferred_numfmts, classification_per_col).

    Only fills columns the user did NOT specify via --numfmt.
    inferred_numfmts: dict {col_letter: fmt_string}
    classification:   dict {col_letter: cls_key}  (used for coercion)
    """
    from openpyxl.utils import get_column_letter
    inferred = {}
    cls_map = {}
    ncol = len(headers)
    for ci in range(ncol):
        col = get_column_letter(ci + 1)
        if col in user_numfmts:
            continue
        samples = []
        all_rows = list(rows) + ([totals] if totals else [])
        for row in all_rows:
            if ci < len(row):
                v = (row[ci] or "").strip() if isinstance(row[ci], str) else ""
                # Skip empties, formulas, and explicit text cells
                if v and not v.startswith("="):
                    samples.append(v)
        cls = _classify_column(samples)
        if cls:
            cls_map[col] = cls
            inferred[col] = _CLS_TO_FMT[cls]
    return inferred, cls_map


def _display_len(s):
    """Width in terminal columns. CJK / full-width chars count as 2."""
    n = 0
    for ch in str(s or ""):
        n += 2 if ord(ch) > 0x4DFF else 1
    return n


def _auto_widths(headers, rows, totals, user_widths):
    """Auto-fit column widths for columns the user did not set via --width.

    Width = max(header_len, max data_len) + 2, clamped to [6, 50].
    """
    from openpyxl.utils import get_column_letter
    out = {}
    ncol = len(headers)
    all_rows = list(rows) + ([totals] if totals else [])
    for ci in range(ncol):
        col = get_column_letter(ci + 1)
        if col in user_widths:
            continue
        max_len = _display_len(headers[ci])
        for row in all_rows:
            if ci < len(row):
                max_len = max(max_len, _display_len(row[ci]))
        out[col] = max(6.0, min(50.0, float(max_len + 2)))
    return out


def _maybe_remove_placeholder(wb):
    """Remove the placeholder sheet created by `new-book` if it is still
    pristine. The placeholder is needed because openpyxl refuses to save a
    workbook with zero sheets, but the model never sees it after the first
    real add-* command."""
    for ws in list(wb.worksheets):
        if (ws.title == "_placeholder"
                and ws.max_row == 1 and ws.max_column == 1
                and ws.cell(1, 1).value is None):
            wb.remove(ws)
            break


def _ensure_unique_sheet_name(wb, name, *, strict=False):
    """If `name` collides, return name with numeric suffix (e.g. 'Q3_2').

    When strict=True, a collision is a hard error. The error message includes
    the existing sheet's real shape (header_cols / data_rows) so the model can
    notice "my previous add-sheet already succeeded" instead of retrying.
    Recovery suggestions stick to pure-CLI args (add-row / set-cell); no
    inline JSON.
    """
    if name not in wb.sheetnames:
        return name
    if strict:
        # Compute existing sheet's real shape so the model knows whether
        # its previous add-sheet already wrote the data.
        ws_existing = wb[name]
        try:
            hdr_cols = _detect_header_cols(ws_existing)
        except Exception:
            hdr_cols = ws_existing.max_column or 0
        last_row = ws_existing.max_row or 0
        data_rows = max(0, last_row - 1) if hdr_cols > 0 else last_row
        raise ValueError(
            f"sheet '{name}' already exists "
            f"(current shape: header_cols={hdr_cols}, data_rows={data_rows}, "
            f"last_row={last_row}).\n"
            f"  If your previous add-sheet/add-table for '{name}' SUCCEEDED, "
            f"the data is already there -- DO NOT rewrite, move to next step.\n"
            f"  To APPEND more rows (no JSON, plain CSV positional args):\n"
            f"    add-row FILE '{name}' \"v1,v2,v3,...\"\n"
            f"  To OVERWRITE one cell:\n"
            f"    set-cell FILE '{name}' B5 \"value\"\n"
            f"  To FORCE a new sheet '{name}_2', re-run add-sheet with "
            f"--allow-dup-rename."
        )
    i = 2
    while f"{name}_{i}" in wb.sheetnames:
        i += 1
    return f"{name}_{i}"


# [C] Normalize repeated/aliased row arguments. Accepts either:
#   --row "a,b,c" --row "d,e,f"   (canonical, preferred)
#   --rows "a,b,c|d,e,f"          (pipe-separated, alias)
#   --rows "a,b,c;d,e,f"          (semicolon-separated, alias)
# and any mix. Returns a flat list of row strings (still CSV inside each).
# Observed in log2.txt: the LLM produced --rows "...|...|..." and
# --rows "...;...;..." three times in a row before giving up. Tolerating
# the alias costs us 6 lines of code and eliminates those retries.
_ROWS_ALIAS_SPLITTERS = ("|", ";")


def _split_rows_spec(rows_alias):
    """Split a single --rows string (with '|' or ';' separators) into a list."""
    if not rows_alias:
        return []
    # Find first non-empty splitter present; prefer '|' over ';'.
    sep = None
    for s in _ROWS_ALIAS_SPLITTERS:
        if s in rows_alias:
            sep = s
            break
    if sep is None:
        # Single row encoded in --rows (no separator). Treat as one row.
        return [rows_alias]
    return [r.strip() for r in rows_alias.split(sep) if r.strip()]


def _merge_rows(row_list, rows_alias, positional_rows):
    """Merge canonical --row, alias --rows, and positional row args."""
    merged = list(row_list or [])
    if rows_alias:
        merged.extend(_split_rows_spec(rows_alias))
    if positional_rows:
        merged.extend([str(r) for r in positional_rows if str(r).strip()])
    return merged


# ---------------------------------------------------------------------------
# A1 / A2 / A3 / A4 helpers (added 2026-04 after log2.txt failure analysis)
# ---------------------------------------------------------------------------
def _expand_packed_rows(rows, *, where="row"):
    """A1: positional row tolerance.

    Real-world failure (log2.txt 15:38:00 / 15:40:59): the model squeezed
    multiple data rows into ONE positional argument joined by '|' or ';',
    e.g. ``"1Tb QLC,$27;1Tb TLC,$29;512Gb,$23"``. argparse passes this as
    a single string, so it ends up as ONE 28-cell row instead of four
    rows of 7 cells -- silent data loss.

    Heuristic split: a string is treated as multi-row iff
        - it contains '|' or ';', AND
        - it contains ',' (otherwise it is a single non-CSV cell), AND
        - splitting by that separator yields >= 2 non-empty fragments
          AND every fragment itself contains a ','.
    The last clause guards against legitimate cell values that happen to
    have a stray ';' (e.g. ``"DDR4 16Gb;eTT,$15.00,..."``): if only ONE
    fragment is csv-shaped, we leave the string intact.

    Returns (expanded_list, n_rows_split). Caller emits a single [WARN]
    summarising n_rows_split so the model learns the canonical form.
    """
    if not rows:
        return list(rows or []), 0
    out = []
    n_split = 0
    for r in rows:
        s = str(r)
        if (("|" in s) or (";" in s)) and ("," in s):
            sep = "|" if "|" in s else ";"
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) >= 2 and all(("," in p) for p in parts):
                out.extend(parts)
                n_split += 1
                continue
        out.append(s)
    return out, n_split


def _warn_packed_rows(n_split, *, cmd):
    """Emit a single concise [WARN] when _expand_packed_rows split rows."""
    if n_split <= 0:
        return
    print(
        f"[WARN][POSITIONAL_ROW_DELIMITER] {cmd}: detected {n_split} "
        f"positional argument(s) packed with '|' or ';' as row separator. "
        f"Auto-split into multiple rows."
    )
    print(
        f"  Canonical form (preferred):\n"
        f"    --row \"a,b,c\" --row \"d,e,f\"          (one --row per data row)\n"
        f"    --rows \"a,b,c|d,e,f\"                  (alias, '|' or ';' splits rows)\n"
        f"  Positional form (one row per arg):\n"
        f"    {cmd} FILE NAME HEADERS \"a,b,c\" \"d,e,f\"  (each arg is ONE row)"
    )


def _detect_header_cols(ws, *, max_scan_rows=4):
    """A2 helper: best-effort column count of the sheet's header row.

    add-sheet writes headers as bold + colored fill. We pick the row in
    the first ``max_scan_rows`` with the largest count of bold cells; if
    no bold row is found, we fall back to the row with the most non-empty
    cells. Returns 0 when the sheet has no usable header.
    """
    if ws is None or (ws.max_row or 0) < 1:
        return 0
    best_bold, best_bold_count = 0, 0
    best_any, best_any_count = 0, 0
    upper = min(max_scan_rows, ws.max_row)
    for rr in range(1, upper + 1):
        bold_cnt = 0
        any_cnt = 0
        for cell in ws[rr]:
            if cell.value is None:
                continue
            sv = str(cell.value).strip()
            if not sv:
                continue
            any_cnt += 1
            try:
                if cell.font and cell.font.bold:
                    bold_cnt += 1
            except Exception:
                pass
        if bold_cnt > best_bold_count:
            best_bold, best_bold_count = rr, bold_cnt
        if any_cnt > best_any_count:
            best_any, best_any_count = rr, any_cnt
    if best_bold_count >= 2:
        return best_bold_count
    return best_any_count


def _warn_col_mismatch(cmd, *, expected, got, where):
    """A2: emit [WARN][COL_MISMATCH] without mutating the row."""
    if expected <= 0 or got == expected:
        return
    print(
        f"[WARN][COL_MISMATCH] {cmd} {where}: expected {expected} "
        f"column(s) (per header), got {got}."
    )
    if got < expected:
        print(
            f"  Missing {expected - got} cell(s) -- written as blanks. "
            f"Pad your CSV: \"v1,v2,...,v{expected}\""
        )
    else:
        print(
            f"  Extra {got - expected} cell(s) -- truncated. Drop the "
            f"trailing values or fix --headers to {got} columns."
        )


def _row_signature(values):
    """A3 helper: stable signature for duplicate-row detection."""
    return tuple(("" if v is None else str(v).strip()) for v in values)


def _warn_duplicate_row(cmd, *, prev_idx, cur_idx, signature):
    """A3: emit [WARN][DUPLICATE_ROW]; row is still written."""
    preview = ",".join(list(signature)[:4])
    if len(signature) > 4:
        preview += ",..."
    print(
        f"[WARN][DUPLICATE_ROW] {cmd}: row {cur_idx} is identical to "
        f"row {prev_idx} ({preview}). Wrote anyway."
    )
    print(
        f"  If unintended, fix with:\n"
        f"    python xlsx_helper.py set-cell FILE --sheet S "
        f"--cell A{cur_idx} --value \"\"   (clear)\n"
        f"  Otherwise, ignore this warning."
    )


def _print_sheet_state(ws, *, sheet_name, header_row_hint=None):
    """A4: print a [STATE] summary so the model knows the real shape.

    The line is single-line and machine-readable; the model can grep it
    to decide when to stop adding rows or move on.
    """
    if ws is None:
        return
    last_row = ws.max_row or 0
    header_cols = _detect_header_cols(ws)
    # Header row: caller may pass a hint (add-sheet knows it exactly);
    # otherwise infer it as the row matching header_cols best.
    h_row = header_row_hint
    if h_row is None and header_cols > 0:
        for rr in range(1, min(5, last_row + 1)):
            cnt = sum(1 for c in ws[rr]
                      if c.value is not None and str(c.value).strip())
            if cnt == header_cols:
                h_row = rr
                break
    data_rows = max(0, last_row - (h_row or 0)) if h_row else 0
    print(
        f"[STATE] sheet='{sheet_name}' header_row={h_row or '?'} "
        f"header_cols={header_cols} data_rows={data_rows} "
        f"last_row={last_row}"
    )


def _commit_sheet(wb, args, kind, meta, *, sheet_name=None,
                  header_row_hint=None):
    """Save workbook after one add-* and print [OK] / NEXT_STEP / [STATE]."""
    meta["sheet_count"] = len(wb.sheetnames)
    _meta_save(wb, meta)
    wb.save(args.file)
    n = meta["sheet_count"]
    extra = f" '{sheet_name}'" if sheet_name else ""
    print(f"[OK] {kind}{extra} -> {args.file} "
          f"(sheets={n}, auto_color={meta.get('auto_color', False)})")
    # A4: emit [STATE] for the sheet we just touched so the model can see
    # the real shape (header_cols, data_rows) and stop guessing.
    if sheet_name and sheet_name in wb.sheetnames:
        _print_sheet_state(
            wb[sheet_name],
            sheet_name=sheet_name,
            header_row_hint=header_row_hint,
        )
    _next_step([
        f"Sheet count = {n}. Continue with another add-* "
        f"(add-sheet / add-assumptions / add-dashboard)",
        f"or finish: python xlsx_helper.py inspect {args.file}",
    ])


def _apply_numfmts(ws, numfmts):
    """Apply per-column number formats to every populated cell in that col."""
    for col, fmt in numfmts.items():
        try:
            col_idx = column_index_from_string(col)
        except Exception:
            print(f"[WARN] numfmt col '{col}' invalid, ignored")
            continue
        for rr in range(1, ws.max_row + 1):
            c = ws.cell(rr, col_idx)
            if c.value is None:
                continue
            # Apply to numbers and to formulas (Excel will compute the value
            # and the format applies to the result).
            if isinstance(c.value, (int, float)) or _is_formula(c.value):
                c.number_format = fmt


def _apply_widths(ws, widths):
    """Apply per-column widths (already validated to numeric)."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _apply_auto_color_block(ws, header_row, meta):
    """Apply blue/black financial coloring to the data block (post-headers)."""
    if not meta.get("auto_color"):
        return
    start = (header_row + 1) if header_row is not None else 1
    for row_cells in ws.iter_rows(min_row=start, values_only=False):
        for cell in row_cells:
            if cell.value is None:
                continue
            _apply_auto_color(cell)


# --- new-book --------------------------------------------------------------
def cmd_new_book(args):
    """Initialize an empty workbook. The `--auto-color` flag is persisted in
    metadata and inherited by every subsequent add-* command."""
    wb = Workbook()
    # Drop the default 'Sheet' so the model's first add-sheet starts clean.
    if wb.active and wb.active.title == "Sheet" and len(wb.sheetnames) == 1:
        wb.remove(wb.active)
    if not wb.sheetnames:
        # openpyxl refuses to save a sheet-less workbook -- park a sentinel
        # placeholder; first add-* call removes it transparently.
        wb.create_sheet(title="_placeholder")
    auto_color = bool(getattr(args, "auto_color", False))
    _meta_save(wb, {"sheet_count": 0, "auto_color": auto_color})
    wb.save(args.file)
    print(f"[OK] new-book -> {args.file} "
          f"(sheets=0, auto_color={auto_color})")
    _next_step([
        "Add sheets one at a time (each command appends one sheet):",
        f"  python xlsx_helper.py add-sheet {args.file} \\",
        "      --name    \"Q3 Revenue\" --title \"Q3 2025 Revenue\" \\",
        "      --headers \"Region,Jan,Feb,Mar,Q1\" \\",
        "      --row     \"NA,100,110,120,=SUM(B4:D4)\" \\",
        "      --row     \"EU,80,85,90,=SUM(B5:D5)\" \\",
        "      --totals  \"Total,=SUM(B4:B5),=SUM(C4:C5),=SUM(D4:D5),=SUM(E4:E5)\"",
        f"  python xlsx_helper.py add-assumptions {args.file} "
        "--kv \"Growth;0.18;Plan\"",
        f"  python xlsx_helper.py add-dashboard {args.file} "
        "--kpi \"=SUM(...);Total\"",
    ])


# --- add-sheet -------------------------------------------------------------
def cmd_add_sheet(args):
    """Append a generic data sheet: title + headers + rows + optional totals.

    Cells starting with '=' are treated as Excel formulas; non-formula
    strings that look numeric are coerced to int/float so number formats
    apply correctly.

    Argument handling is LLM-friendly:
      * `--name` / `--headers` / `--row ...`          (canonical)
      * positional fallback: `add-sheet FILE NAME HEADERS [ROW1 ROW2 ...]`
      * `--rows "a,b,c|d,e,f"` alias (pipe or semicolon row separator)
    See log2.txt 11:07:12 - 11:08:16 for the five failure forms this fixes.
    """
    wb = _open_book(args.file)
    meta = _meta_load(wb)
    _maybe_remove_placeholder(wb)

    # [C] Positional-spill correction. argparse fills name_pos/headers_pos
    # greedily left-to-right even when --name/--headers were already given
    # by keyword. That would shadow real row data. Detect shadowing and
    # push the "extra" positionals down into rows_pos so they become data.
    spilled_into_rows = []
    if args.name and getattr(args, "name_pos", None):
        spilled_into_rows.append(args.name_pos)
        args.name_pos = None
    if args.headers and getattr(args, "headers_pos", None):
        spilled_into_rows.append(args.headers_pos)
        args.headers_pos = None
    if spilled_into_rows:
        existing = list(getattr(args, "rows_pos", None) or [])
        args.rows_pos = spilled_into_rows + existing

    # [C] Merge canonical (--name) and positional (name_pos) into one value.
    name_val = args.name or getattr(args, "name_pos", None)
    if not name_val or not str(name_val).strip():
        raise ValueError(
            "add-sheet: missing sheet name. Use one of:\n"
            "  add-sheet FILE.xlsx --name \"Q3 Revenue\" --headers \"...\" --row \"...\"\n"
            "  add-sheet FILE.xlsx \"Q3 Revenue\" \"Region,Jan,Feb\" "
            "\"NA,100,110\" \"EU,80,85\""
        )
    name = _clip_str(str(name_val), _LIM["sheet_name"], where="sheet name")
    # [D] Default to strict rename: duplicates become hard errors unless
    # the caller explicitly passed --allow-dup-rename.
    strict_dup = not bool(getattr(args, "allow_dup_rename", False))
    name = _ensure_unique_sheet_name(wb, name, strict=strict_dup)

    # [C] Merge canonical (--headers) and positional (headers_pos).
    headers_val = args.headers or getattr(args, "headers_pos", None)
    if not headers_val or not str(headers_val).strip():
        raise ValueError(
            "add-sheet: missing headers. Use one of:\n"
            "  add-sheet FILE.xlsx --name \"S\" --headers \"a,b,c\" --row \"1,2,3\"\n"
            "  add-sheet FILE.xlsx \"S\" \"a,b,c\" \"1,2,3\" \"4,5,6\""
        )
    headers = _split_csv(str(headers_val))
    headers = _clip_list(headers, _LIM["headers_max"], where="headers")
    headers = [_clip_str(h, _LIM["cell"], where=f"header[{i}]")
               for i, h in enumerate(headers)]
    if not headers:
        raise ValueError("add-sheet: --headers parsed to empty list")

    # [C] Merge --row (repeatable), --rows (alias with '|'/';' separator),
    # and positional row_pos (anything left after file/name/headers).
    # [A1] Before merging, detect packed positional rows (single arg with
    # '|' or ';' that hides multiple csv rows). Auto-split + warn so the
    # model learns the canonical form for the next call.
    pos_rows_in = list(getattr(args, "rows_pos", None) or [])
    pos_rows_expanded, n_split = _expand_packed_rows(pos_rows_in,
                                                     where="add-sheet rows")
    _warn_packed_rows(n_split, cmd="add-sheet")
    rows_raw = _merge_rows(
        args.row,
        getattr(args, "rows_alias", None),
        pos_rows_expanded,
    )
    rows_raw = _clip_list(rows_raw, _LIM["rows_max"], where="rows")
    rows = []
    raw_signatures = []  # for A3 duplicate detection
    for ri, row_str in enumerate(rows_raw):
        cells = _split_csv(row_str)
        # [A2] Warn when this row's column count != len(headers) BEFORE
        # padding/truncating. Helps the model spot inconsistent --headers
        # vs row width on the very next step.
        _warn_col_mismatch(
            "add-sheet",
            expected=len(headers),
            got=len(cells),
            where=f"row[{ri + 1}]",
        )
        if len(cells) < len(headers):
            cells = cells + [""] * (len(headers) - len(cells))
        elif len(cells) > len(headers):
            cells = cells[: len(headers)]
        cells = [_clip_str(c, _LIM["cell"], where=f"row[{ri}].col[{ci}]")
                 for ci, c in enumerate(cells)]
        # [A3] Compare against previous row signature; flag exact dup.
        sig = _row_signature(cells)
        if raw_signatures and sig == raw_signatures[-1] and any(sig):
            _warn_duplicate_row(
                "add-sheet", prev_idx=ri, cur_idx=ri + 1, signature=sig,
            )
        raw_signatures.append(sig)
        rows.append(cells)

    totals = []
    if args.totals:
        totals = _split_csv(args.totals)
        if len(totals) < len(headers):
            totals = totals + [""] * (len(headers) - len(totals))
        elif len(totals) > len(headers):
            totals = totals[: len(headers)]
        totals = [_clip_str(t, _LIM["cell"], where=f"totals[{ci}]")
                  for ci, t in enumerate(totals)]

    # Parse --numfmt and --width specs (both repeatable).
    numfmts = {}
    for spec in (args.numfmt or []):
        col, val = _parse_col_pair(spec)
        if col and val:
            numfmts[col] = val
        else:
            print(f"[WARN] --numfmt '{spec}' ignored "
                  "(use 'COL:FMT' e.g. 'B:#,##0')")
    widths = {}
    for spec in (args.width or []):
        col, val = _parse_col_pair(spec)
        if not (col and val):
            print(f"[WARN] --width '{spec}' ignored "
                  "(use 'COL:N' e.g. 'A:18')")
            continue
        try:
            widths[col] = float(val)
        except ValueError:
            print(f"[WARN] --width '{spec}' value not numeric, ignored")

    title = _clip_str(args.title, _LIM["title"], where="title") \
        if args.title else ""
    header_fill = (args.header_fill or "1F3864").lstrip("#")
    freeze = not bool(args.no_freeze_top)

    # [C-plan] Smart defaults: infer per-column numfmt + column width from
    # the raw data so the model rarely needs --numfmt/--width. Explicit
    # flags still win; inference only fills the gaps.
    inferred_numfmts, col_cls = _auto_infer_numfmts(
        headers, rows, totals, numfmts)
    merged_numfmts = dict(inferred_numfmts)
    merged_numfmts.update(numfmts)  # user explicit wins over inferred

    inferred_widths = _auto_widths(headers, rows, totals, widths)
    merged_widths = dict(inferred_widths)
    merged_widths.update(widths)  # user explicit wins

    ws = wb.create_sheet(title=name)
    r = 1
    if title:
        ws.cell(r, 1).value = title
        ws.cell(r, 1).font = Font(bold=True, size=14)
        r += 1
        r += 1  # blank spacer row

    header_row = r
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(r, ci)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    # Build col-letter lookup once so we can pick the right coercion per col.
    from openpyxl.utils import get_column_letter as _gcl
    col_letters = [_gcl(ci + 1) for ci in range(len(headers))]

    for row_data in rows:
        for ci, raw in enumerate(row_data, start=1):
            if raw == "":
                continue
            cell = ws.cell(r, ci)
            cls = col_cls.get(col_letters[ci - 1])
            cell.value = _coerce_with_classification(raw, cls)
        r += 1

    if totals:
        for ci, raw in enumerate(totals, start=1):
            if raw == "":
                continue
            cell = ws.cell(r, ci)
            cls = col_cls.get(col_letters[ci - 1])
            cell.value = _coerce_with_classification(raw, cls)
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style="thin"))
        r += 1

    _apply_widths(ws, merged_widths)
    _apply_numfmts(ws, merged_numfmts)

    if freeze:
        ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate

    _apply_auto_color_block(ws, header_row, meta)
    _commit_sheet(wb, args, "add-sheet", meta, sheet_name=name,
                  header_row_hint=header_row)
    # [C] Loud warning when no data rows were provided. The silent "headers
    # only" sheet is the #1 root cause of the Flash-report bug in log2.txt:
    # the model called add-sheet 3 times for 'Flash报价' without --row and
    # never noticed the sheet had no body. A loud block in stdout makes the
    # mistake visible in the very next step's returnData.
    if not rows:
        print("\n" + "=" * 64)
        print("[WARN][NO_DATA_ROWS] Sheet '{}' has HEADERS ONLY.".format(name))
        print("  You did NOT pass any --row / --rows / positional row args.")
        print("  Fix options:")
        print("    (A) Re-run add-sheet with --row \"v1,v2,v3\" --row ... ")
        print("        (requires --allow-dup-rename or delete the sheet first)")
        print("    (B) Append data rows to the existing sheet (canonical):")
        print("        python xlsx_helper.py add-row FILE.xlsx '{}' "
              "v1 v2 v3 ...".format(name))
        print("    (C) Bulk write a rectangular 2D area (CSV per ROW):")
        print("        python xlsx_helper.py set-range FILE.xlsx '{}' A2 "
              "\"v1,v2,v3\" \"v4,v5,v6\"".format(name))
        print("=" * 64)


# --- add-row (canonical) ---------------------------------------------------
def cmd_add_row(args):
    """Append a row (default) or overwrite a row at --at IDX.

    Canonical name (matches the add-sheet/add-assumptions/add-dashboard
    family). The LLM spontaneously used `add-row` in log3.txt and the old
    helper rejected it. cmd_add_row reuses cmd_write_row by injecting
    row_idx=0 (append) unless --at was passed.
    """
    # [1A] Map add-row's --at IDX onto write-row's positional row_idx.
    args.row_idx = args.at_idx if getattr(args, "at_idx", None) else "0"
    return cmd_write_row(args, _via_alias="add-row")


# --- write-row (deprecated alias of add-row) -------------------------------
def cmd_write_row(args, *, _via_alias=None):
    """Append/replace a single row in an existing sheet.

    Calling semantics:
      * row_idx = 0   -> append after the last populated row
      * row_idx >= 1  -> write (or overwrite) at that 1-based row
      * values        -> either CSV in a single arg, or many positional args
    Everything starting with '=' is treated as a formula.

    Subcommand `write-row` is kept as a deprecated alias of `add-row`
    (canonical). Pre-1A logs show the model used both names; we accept
    either but emit a one-line [DEPRECATED] notice on the alias.
    """
    if _via_alias is None:
        # Direct write-row invocation: warn the model to migrate to add-row.
        print(
            "[DEPRECATED] `write-row` is now an alias of `add-row`. "
            "Prefer:\n"
            "  python xlsx_helper.py add-row FILE.xlsx SHEET v1 v2 ...   "
            "(append)\n"
            "  python xlsx_helper.py add-row FILE.xlsx SHEET v1 v2 ... --at IDX  "
            "(overwrite row IDX)"
        )
    wb = _open_book(args.file)
    label = _via_alias or "write-row"
    if args.sheet not in wb.sheetnames:
        raise ValueError(
            f"{label}: sheet '{args.sheet}' not found. "
            f"Existing sheets: {wb.sheetnames}"
        )
    ws = wb[args.sheet]
    # [1A] Row values can come in two shapes:
    #   add-row FILE SHEET "a,b,c"         (one CSV string)
    #   add-row FILE SHEET a b c           (many positional)
    vals = list(args.values or [])
    if len(vals) == 1 and ("," in vals[0]):
        vals = _split_csv(vals[0])
    label = _via_alias or "write-row"
    if not vals:
        raise ValueError(
            f"{label}: no values provided. Usage:\n"
            f"  add-row FILE SHEET v1 v2 v3 ...                (append)\n"
            f"  add-row FILE SHEET v1 v2 v3 ... --at IDX       (overwrite row IDX)\n"
            f"  add-row FILE SHEET \"v1,v2,v3\"                  (CSV-in-one)"
        )
    try:
        row_idx = int(args.row_idx)
    except (TypeError, ValueError):
        raise ValueError(f"{label}: row index must be an integer, got '{args.row_idx}'")
    if row_idx == 0:
        # Append after last non-empty row.
        row_idx = (ws.max_row or 0) + 1
    if row_idx < 1:
        raise ValueError(f"{label}: row index must be >= 1 (or 0 to append), got {row_idx}")

    # [A2] Compare incoming column count to the sheet's existing header
    # row so the model is alerted when headers and rows drift out of sync.
    header_cols = _detect_header_cols(ws)
    _warn_col_mismatch(
        label, expected=header_cols, got=len(vals),
        where=f"row {row_idx}",
    )

    # [A3] Detect "I just wrote the same row twice in a row". Compare the
    # incoming values against what is currently at row_idx-1 (only when
    # appending to a non-empty sheet). Cheap, no hashing of the whole sheet.
    if row_idx > 1:
        prev_vals = []
        for ci in range(1, max(1, len(vals)) + 1):
            prev_vals.append(ws.cell(row_idx - 1, ci).value)
        prev_sig = _row_signature(prev_vals)
        cur_sig = _row_signature(vals)
        if any(cur_sig) and prev_sig == cur_sig:
            _warn_duplicate_row(
                label, prev_idx=row_idx - 1, cur_idx=row_idx,
                signature=cur_sig,
            )

    for ci, raw in enumerate(vals, start=1):
        raw_clipped = _clip_str(str(raw), _LIM["cell"],
                                where=f"{label} col[{ci}]")
        ws.cell(row_idx, ci).value = _coerce_scalar(raw_clipped)
    wb.save(args.file)
    print(f"[OK] {label} -> {args.file} (sheet='{args.sheet}' "
          f"row={row_idx} cols={len(vals)})")
    # [A4] Print updated [STATE] so the model knows current data_rows
    # without needing a separate inspect call.
    _print_sheet_state(ws, sheet_name=args.sheet)
    _next_step([
        f"Row {row_idx} written with {len(vals)} cells. "
        f"Continue with more add-row / add-sheet / inspect.",
    ])


# --- add-assumptions -------------------------------------------------------
def cmd_add_assumptions(args):
    """Append a 3-column 'Item / Value / Source' sheet for model inputs.

    Designed for cross-sheet references like `=Inputs!$B$4`.
    """
    wb = _open_book(args.file)
    meta = _meta_load(wb)
    _maybe_remove_placeholder(wb)

    name = _clip_str(args.name or "Assumptions", _LIM["sheet_name"],
                     where="sheet name")
    name = _ensure_unique_sheet_name(wb, name)
    title = _clip_str(args.title or "Assumptions", _LIM["title"],
                      where="title")
    header_fill = (args.header_fill or "1F3864").lstrip("#")

    kv_raw = list(args.kv or [])
    kv_raw = _clip_list(kv_raw, _LIM["kv_max"], where="--kv items")
    if not kv_raw:
        print("[WARN] add-assumptions: no --kv provided; nothing added")
        return

    rows = []
    for i, spec in enumerate(kv_raw):
        # Format: "Item;Value;Source" (Source optional).
        parts = [p.strip() for p in str(spec).split(";")]
        while len(parts) < 3:
            parts.append("")
        parts = parts[:3]
        parts = [_clip_str(p, _LIM["cell"],
                           where=f"--kv[{i}].field[{j}]")
                 for j, p in enumerate(parts)]
        rows.append(parts)

    numfmts = {}
    for spec in (args.numfmt or []):
        col, val = _parse_col_pair(spec)
        if col and val:
            numfmts[col] = val

    ws = wb.create_sheet(title=name)
    r = 1
    ws.cell(r, 1).value = title
    ws.cell(r, 1).font = Font(bold=True, size=14)
    r += 1
    r += 1

    header_row = r
    for ci, h in enumerate(["Item", "Value", "Source"], start=1):
        cell = ws.cell(r, ci)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    for kv in rows:
        for ci, raw in enumerate(kv, start=1):
            if raw == "":
                continue
            cell = ws.cell(r, ci)
            cell.value = _coerce_scalar(raw)
        r += 1

    # Sane defaults for an Inputs-style sheet.
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28

    _apply_numfmts(ws, numfmts)
    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    _apply_auto_color_block(ws, header_row, meta)
    _commit_sheet(wb, args, "add-assumptions", meta, sheet_name=name)


# --- add-dashboard ---------------------------------------------------------
def cmd_add_dashboard(args):
    """Append a 2-column 'Metric / Value' KPI dashboard sheet.

    Each --kpi spec is 'Value_or_Formula;Label' so that putting a formula
    first does not collide with shell-quoting around equals signs.
    """
    wb = _open_book(args.file)
    meta = _meta_load(wb)
    _maybe_remove_placeholder(wb)

    name = _clip_str(args.name or "Dashboard", _LIM["sheet_name"],
                     where="sheet name")
    name = _ensure_unique_sheet_name(wb, name)
    title = _clip_str(args.title or "Dashboard", _LIM["title"],
                      where="title")
    header_fill = (args.header_fill or "1F3864").lstrip("#")

    kpi_raw = list(args.kpi or [])
    kpi_raw = _clip_list(kpi_raw, _LIM["kpi_max"], where="--kpi items")
    if not kpi_raw:
        print("[WARN] add-dashboard: no --kpi provided; nothing added")
        return

    rows = []
    for i, spec in enumerate(kpi_raw):
        parts = [p.strip() for p in str(spec).split(";")]
        while len(parts) < 2:
            parts.append("")
        value, label = parts[0], parts[1]
        value = _clip_str(value, _LIM["cell"], where=f"--kpi[{i}].value")
        label = _clip_str(label, _LIM["cell"], where=f"--kpi[{i}].label")
        rows.append([label, value])

    ws = wb.create_sheet(title=name)
    r = 1
    ws.cell(r, 1).value = title
    ws.cell(r, 1).font = Font(bold=True, size=16)
    r += 1
    r += 1

    header_row = r
    for ci, h in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(r, ci)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    for label, value in rows:
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).value = _coerce_scalar(value)
        ws.cell(r, 2).font = Font(size=14, bold=True)
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    _apply_auto_color_block(ws, header_row, meta)
    _commit_sheet(wb, args, "add-dashboard", meta, sheet_name=name)


# --- catalog ---------------------------------------------------------------
def cmd_catalog(args):
    """Print add-* cookbook (run ONCE per task)."""
    L = []
    L.append("================================================================")
    L.append(" xlsx_helper.py catalog -- copy-pastable command cookbook       ")
    L.append("================================================================")
    L.append("")
    # [C] COMMON MISTAKES block: placed first so models see the usual
    # argument-format pitfalls before the canonical examples. The parser
    # auto-recovers (positional fallback + --rows alias + parse_known_args)
    # but every recovery costs a [WARN] line and confuses downstream steps,
    # so the canonical form is still strongly preferred.
    # === RULE 0 ================================================
    L.append("=== RULE 0: ONE TASK = ONE WORKBOOK FILENAME ===")
    L.append("  Pick the .xlsx path ONCE in new-book, then reuse it in every")
    L.append("  add-*/add-row/set-* call. Switching filenames mid-task leaves")
    L.append("  orphan files and breaks the pipeline.")
    L.append("")
    # === Minimal canonical form =================================
    L.append("=== STEP 1: init an empty workbook ===")
    L.append("  python xlsx_helper.py new-book FILE.xlsx")
    L.append("")
    L.append("=== STEP 2: append one sheet per command (positional form) ===")
    L.append("")
    L.append("  --- add-sheet FILE NAME HEADERS ROW1 ROW2 ... ----------------")
    L.append("  Canonical 4-positional form. Column number formats ($, %, int,")
    L.append("  decimal) and widths are AUTO-DETECTED from your data.")
    L.append("    python xlsx_helper.py add-sheet FILE.xlsx \\")
    L.append('      "Q3 Revenue" \\')
    L.append('      "Region,Jan,Feb,Mar,Q1" \\')
    L.append('      "NA,100,110,120,=SUM(B2:D2)" \\')
    L.append('      "EU,80,85,90,=SUM(B3:D3)" \\')
    L.append('      --totals "Total,=SUM(B2:B3),=SUM(C2:C3),=SUM(D2:D3),=SUM(E2:E3)"')
    L.append("")
    L.append("  --- add-row FILE SHEET v1 v2 ... -----------------------------")
    L.append("  Append (default) or overwrite (--at IDX) one row in an")
    L.append("  existing sheet.")
    L.append("    python xlsx_helper.py add-row FILE.xlsx 'Q3 Revenue' \\")
    L.append('      "Asia,60,70,75,=SUM(B4:D4)"')
    L.append("")
    L.append("  --- add-assumptions FILE ------------------------------------")
    L.append("  3-col Item/Value/Source sheet for inputs referenced as")
    L.append("  '=Inputs!$B$2' by other sheets.")
    L.append("    python xlsx_helper.py add-assumptions FILE.xlsx \\")
    L.append('      --name "Inputs" \\')
    L.append('      --kv "Growth;0.18;Q3 Board Memo" \\')
    L.append('      --kv "Op Margin;0.15;FY2026 Plan"')
    L.append("")
    L.append("  --- add-dashboard FILE --------------------------------------")
    L.append("  KPI cards (2 cols: Metric / Value). Cross-sheet formulas OK.")
    L.append("    python xlsx_helper.py add-dashboard FILE.xlsx \\")
    L.append('      --name "Summary" \\')
    L.append('      --kpi "=SUM(\'Q3 Revenue\'!E2:E3);Total Revenue"')
    L.append("")
    L.append("=== STEP 3: inspect / edit =====================================")
    L.append("  python xlsx_helper.py inspect  FILE.xlsx")
    L.append("  python xlsx_helper.py extract  FILE.xlsx")
    L.append("  # set-cell FILE SHEET A1 VALUE  (VALUE auto-detect: '=' ->")
    L.append("  # formula, number -> number, else text)")
    L.append("  python xlsx_helper.py set-cell FILE.xlsx 'Q3 Revenue' B4 120")
    L.append("  python xlsx_helper.py set-cell FILE.xlsx 'Q3 Revenue' F4 "
             "\"=SUM(B4:E4)\"")
    L.append("  # set-range FILE SHEET A1 \"v1,v2,..\" \"v3,v4,..\"  (CSV per ROW)")
    L.append("  python xlsx_helper.py set-range FILE.xlsx 'Q3 Revenue' B2 "
             "\"100,110,120\" \"80,85,90\"")
    L.append("")
    # === Common pitfalls =========================================
    L.append("=== COMMON MISTAKES ===")
    L.append("  WRONG: switching filename mid-task  (data.xlsx -> prices.xlsx)")
    L.append("  RIGHT: reuse the SAME path from new-book in every later call.")
    L.append("  WRONG: add-sheet FILE --name S --headers H  (NO row data!)")
    L.append("  RIGHT: always pass rows too, or follow with add-row. Headers-")
    L.append("         only sheets emit a loud [WARN][NO_DATA_ROWS] block.")
    L.append("  WRONG: add-sheet with --name of an existing sheet  (dup)")
    L.append("  RIGHT: amend existing sheets with set-cell / set-range / add-row.")
    L.append("         --allow-dup-rename forces _2 suffix only when you really")
    L.append("         want a fresh sheet.")
    L.append("")
    L.append("  --- Vocabulary aliases (no need to memorize two names) ---")
    L.append("  add-table   == add-sheet   (xlsx sheets ARE tables)")
    L.append("  write-row   == add-row     (deprecated; add-row is canonical)")
    L.append("")
    # === Layout convention =======================================
    L.append("=== LAYOUT CONVENTION (needed for writing formulas) ===")
    L.append("  if no --title:   row 1 = headers, row 2+ = data, then totals")
    L.append("  if --title T:    row 1 = title, row 2 blank, row 3 = headers,")
    L.append("                   row 4+ = data, then totals")
    L.append("")
    # === Advanced (optional) =====================================
    L.append("=== ADVANCED (optional) -- use only when autofit is wrong ===")
    L.append("  --numfmt B:0.0%     explicit column format (overrides autofmt)")
    L.append("  --width  A:18       explicit column width (overrides autofit)")
    L.append("  --title  \"...\"      big title row above the header row")
    L.append("  --header-fill 2E75B6    6-hex header background color")
    L.append("  --no-freeze-top     skip the freeze_panes on header row")
    L.append("")
    L.append("=== Hard caps (auto-truncate, never fail) ===")
    L.append(f"  sheet name <={_LIM['sheet_name']} (Excel limit)    "
             f"title <={_LIM['title']}    cell <={_LIM['cell']}")
    L.append(f"  headers <={_LIM['headers_max']} cols    "
             f"rows <={_LIM['rows_max']} per add-sheet")
    L.append(f"  --kv <={_LIM['kv_max']} per add-assumptions    "
             f"--kpi <={_LIM['kpi_max']} per add-dashboard")
    L.append("")
    L.append("=== Formula rule (financial-model bedrock) ===")
    L.append("  ALWAYS write Excel formulas (=SUM, =B4*C4, =Sheet!$B$4),")
    L.append("  NEVER pre-compute in Python and hardcode the result. That ")
    L.append("  kills the spreadsheet's recalculation behavior.")
    print("\n".join(L))
    _next_step([
        "Now run new-book, then append one sheet per add-* command:",
        "  python xlsx_helper.py new-book ${WORKSPACE}/book.xlsx --auto-color",
        "  python xlsx_helper.py add-sheet ${WORKSPACE}/book.xlsx "
        "--name \"...\" --headers \"...\" --row \"...\"",
        "  python xlsx_helper.py add-assumptions / add-dashboard / ...",
    ])


# ===========================================================================
# Reference manual the model reads after every command. We just dump the
# sibling SKILL.md (frontmatter + first H1 stripped) so there is a SINGLE
# source of truth for the CLI surface; updating SKILL.md updates every
# command's tail-output.
# ===========================================================================
_SKILL_MD_BODY_CACHE = None


def _load_skill_md_body():
    """Read sibling SKILL.md, strip YAML frontmatter and the first H1 title.
    Cached for the lifetime of the process. Returns "" on any IO error."""
    global _SKILL_MD_BODY_CACHE
    if _SKILL_MD_BODY_CACHE is not None:
        return _SKILL_MD_BODY_CACHE
    body = ""
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        md_path = os.path.normpath(os.path.join(here, os.pardir, "SKILL.md"))
        with open(md_path, "r", encoding="utf-8") as f:
            text = f.read()
        lines = text.splitlines()
        i, n = 0, len(lines)
        # 1) skip YAML frontmatter delimited by leading `---`
        if i < n and lines[i].strip() == "---":
            i += 1
            while i < n and lines[i].strip() != "---":
                i += 1
            if i < n:
                i += 1
        # 2) skip blank lines after frontmatter
        while i < n and not lines[i].strip():
            i += 1
        # 3) skip the first H1 title line (e.g. "# Xlsx-Editor")
        if i < n and lines[i].lstrip().startswith("# "):
            i += 1
        # 4) skip blanks after the title
        while i < n and not lines[i].strip():
            i += 1
        body = "\n".join(lines[i:]).rstrip() + "\n"
    except Exception:
        body = ""
    _SKILL_MD_BODY_CACHE = body
    return body


def _print_commands_menu(stream=None):
    """Print SKILL.md (without frontmatter/title) so the model always has the
    full skill manual at the tail of every command's output. Called on both
    success and failure paths."""
    if stream is None:
        stream = sys.stdout
    body = _load_skill_md_body()
    stream.write("\n=== SKILL.md ===\n")
    if body:
        stream.write(body)
    else:
        # Fallback: SKILL.md missing/unreadable; print a one-line escape hatch
        # so the model still knows where the canonical manual lives.
        stream.write("(SKILL.md not found beside this script; "
                     "open skills/xlsx-editor/SKILL.md for the manual)\n")


# Hallucinated subcommand -> intent-aware "did you mean" hint. Argparse's
# default invalid-choice error already lists every legal cmd, but the model
# still picks the wrong one when its naming intuition does not match (e.g.
# write-range vs set-range, create_table vs add-sheet). Map known wrong
# names to the canonical replacement so the model sees a directed fix.
_XLSX_DID_YOU_MEAN = {
    "create_table":  "add-sheet FILE.xlsx \"Sheet\" \"a,b,c\" \"1,2,3\" \"4,5,6\"  "
                     "(sheet with headers + data rows in one shot)",
    "create-table":  "add-sheet FILE.xlsx \"Sheet\" \"a,b,c\" \"1,2,3\" \"4,5,6\"",
    "add-data":      "add-sheet FILE.xlsx \"Sheet\" \"a,b\" \"1,2\"  "
                     "or  add-row FILE.xlsx \"Sheet\" \"v1,v2\"",
    "write-range":   "set-range FILE.xlsx \"Sheet\" A1 \"v1,v2,v3\" [\"v4,v5,v6\" ...]",
    "fill-range":    "set-range FILE.xlsx \"Sheet\" A1 \"v1,v2,v3\" [\"v4,v5,v6\" ...]",
    "update-range":  "set-range FILE.xlsx \"Sheet\" A1 \"v1,v2,v3\" [\"v4,v5,v6\" ...]",
    "write-cell":    "set-cell FILE.xlsx \"Sheet\" A1 VALUE",
    "update-cell":   "set-cell FILE.xlsx \"Sheet\" A1 VALUE",
    "new":               "new-book FILE.xlsx",
    "new-xlsx":          "new-book FILE.xlsx",
    "new-workbook":      "new-book FILE.xlsx",
    "new-spreadsheet":   "new-book FILE.xlsx",
    "new-sheet":     "add-sheet FILE.xlsx \"Sheet\" \"a,b\" \"1,2\"  "
                     "(needs an existing workbook -- run new-book first)",
}


def _intercept_hallucinated_cmd(known_cmds):
    """If sys.argv[1] is a well-known wrong name, print a directed fix and
    exit(2) BEFORE argparse runs. Falls through silently otherwise."""
    if len(sys.argv) < 2:
        return
    cmd = sys.argv[1]
    if cmd in known_cmds or cmd.startswith("-"):
        return
    if cmd in _XLSX_DID_YOU_MEAN:
        sys.stderr.write(
            f"[ERROR] unknown subcommand '{cmd}'. "
            f"Did you mean:  {_XLSX_DID_YOU_MEAN[cmd]}\n"
        )
        _print_commands_menu(stream=sys.stderr)
        sys.exit(2)


class SmartArgumentParser(argparse.ArgumentParser):
    def error(self, message):
        sys.stderr.write(f"[ERROR] {message}\n")
        # Print the full commands menu so the model can pick a different
        # subcommand (or fix args of the same one).
        _print_commands_menu(stream=sys.stderr)
        sys.exit(2)


# ===========================================================================
# Argparse wiring
# ===========================================================================
_XLSX_KNOWN_CMDS = {
    "catalog", "new-book",
    "add-sheet", "add-row", "write-row", "add-table",
    "add-assumptions", "add-dashboard",
    "inspect", "extract", "set-cell", "set-range",
    "new-from-outline", "examples", "create",
}


def main():
    # Intercept well-known wrong subcommand names BEFORE argparse so the model
    # gets a directed fix (e.g. write-range -> set-range, create_table ->
    # add-sheet) instead of a generic "invalid choice" listing.
    _intercept_hallucinated_cmd(_XLSX_KNOWN_CMDS)

    ap = SmartArgumentParser(prog="xlsx_helper",
                             description="Programmatic helper for .xlsx files.")
    sp = ap.add_subparsers(dest="cmd", required=True)

    # -- new incremental block API (preferred path) ----------------------
    p_cat = sp.add_parser(
        "catalog",
        help="Print add-* cookbook (run ONCE at task start)")
    p_cat.set_defaults(func=cmd_catalog)

    p_nb = sp.add_parser(
        "new-book",
        help="Initialize an empty .xlsx (then append sheets via add-*)")
    p_nb.add_argument("file", help="output .xlsx path")
    p_nb.add_argument("--auto-color", action="store_true",
                      help="apply financial color standard "
                           "(blue=hardcoded, black=formula) to every "
                           "subsequent add-* sheet")
    p_nb.set_defaults(func=cmd_new_book)

    p_as = sp.add_parser(
        "add-sheet",
        help="Append a generic data sheet (title+headers+rows[+totals])")
    p_as.add_argument("file")
    # [C] --name / --headers are now OPTIONAL so the LLM-friendly positional
    # fallback can fill them in. cmd_add_sheet validates non-empty.
    p_as.add_argument("--name", default=None,
                      help=f"sheet name (<= {_LIM['sheet_name']} chars)")
    p_as.add_argument("--title", default=None,
                      help=f"sheet title row (<= {_LIM['title']} chars)")
    p_as.add_argument("--headers", default=None,
                      help="comma-separated headers, "
                           f"<= {_LIM['headers_max']} cols")
    p_as.add_argument("--row", action="append", default=[],
                      help="repeatable; one comma-separated data row "
                           f"(<= {_LIM['rows_max']} total)")
    # [C] --rows alias: accepts "a,b,c|d,e,f" or "a,b,c;d,e,f" in a single
    # string. Observed misuse in log2.txt; tolerated but canonical is --row.
    p_as.add_argument("--rows", dest="rows_alias", default=None,
                      help="alias for repeated --row, rows separated by '|' or ';'")
    p_as.add_argument("--totals", default=None,
                      help="optional comma-separated totals row "
                           "(formulas welcome)")
    p_as.add_argument("--numfmt", action="append", default=[],
                      help="repeatable 'COL:FMT' e.g. 'B:#,##0' 'C:0.00%'")
    p_as.add_argument("--width", action="append", default=[],
                      help="repeatable 'COL:N' e.g. 'A:18'")
    p_as.add_argument("--header-fill", default=None,
                      help="hex fill for header band (default 1F3864)")
    p_as.add_argument("--no-freeze-top", action="store_true",
                      help="disable header-row freeze pane")
    # [D] strict dup-name behavior: default ERROR, opt out with this flag.
    p_as.add_argument("--allow-dup-rename", action="store_true",
                      dest="allow_dup_rename",
                      help="if sheet name collides, silently rename to NAME_2 "
                           "(default: hard error with recovery hint)")
    # [C] positional fallback. With nargs and parse_known_args, the CLI
    # accepts: add-sheet FILE NAME HEADERS ROW1 ROW2 ... in any order the
    # model tries first.
    p_as.add_argument("name_pos", nargs="?", default=None,
                      help="positional fallback for --name")
    p_as.add_argument("headers_pos", nargs="?", default=None,
                      help="positional fallback for --headers")
    p_as.add_argument("rows_pos", nargs="*", default=[],
                      help="positional fallback for --row (repeat allowed)")
    p_as.set_defaults(func=cmd_add_sheet)

    # [1A] CANONICAL: `add-row` (matches add-sheet/add-assumptions/add-dashboard
    # naming convention; the LLM spontaneously called this in log3.txt S3).
    # Default semantics: APPEND after the last populated row. Overwrite an
    # existing row by passing `--at IDX`.
    p_ar = sp.add_parser(
        "add-row",
        help="Append a row to an existing sheet (or overwrite via --at IDX)")
    p_ar.add_argument("file")
    p_ar.add_argument("sheet", help="existing sheet name")
    p_ar.add_argument("values", nargs="+",
                      help="cell values (CSV-in-one or multiple positional)")
    p_ar.add_argument("--at", dest="at_idx", default=None,
                      help="1-based row index to overwrite (default: append)")
    p_ar.set_defaults(func=cmd_add_row)

    # [1A] DEPRECATED ALIAS: `write-row FILE SHEET ROW_IDX v1 v2 ...`.
    # Kept for backward compat (older Agent runs may have learned this name).
    # cmd_write_row prints a [DEPRECATED] notice steering the model to add-row.
    p_wr = sp.add_parser(
        "write-row",
        help="[DEPRECATED] alias of add-row (kept for backward compat)")
    p_wr.add_argument("file")
    p_wr.add_argument("sheet", help="existing sheet name")
    p_wr.add_argument("row_idx",
                      help="1-based row index, or 0 to append after last row")
    p_wr.add_argument("values", nargs="+",
                      help="cell values (CSV-in-one or multiple positional)")
    p_wr.set_defaults(func=cmd_write_row)

    # [1B] ALIAS: `add-table` is what docx/pptx call "create a table-shaped
    # block". An xlsx sheet IS a table (header row + data rows + optional
    # totals). Reusing the docx/pptx vocabulary lets the LLM cross-recall
    # without retrying. Observed in log3.txt S7+S8: model used add-table
    # twice on xlsx (each time failing with `invalid choice`).
    p_at_xlsx = sp.add_parser(
        "add-table",
        help="Alias for add-sheet (an xlsx sheet IS a table)")
    p_at_xlsx.add_argument("file")
    p_at_xlsx.add_argument("--name", default=None)
    p_at_xlsx.add_argument("--title", default=None)
    p_at_xlsx.add_argument("--headers", default=None)
    p_at_xlsx.add_argument("--row", action="append", default=[])
    p_at_xlsx.add_argument("--rows", dest="rows_alias", default=None)
    p_at_xlsx.add_argument("--totals", default=None)
    p_at_xlsx.add_argument("--numfmt", action="append", default=[])
    p_at_xlsx.add_argument("--width", action="append", default=[])
    p_at_xlsx.add_argument("--header-fill", default=None)
    p_at_xlsx.add_argument("--no-freeze-top", action="store_true")
    p_at_xlsx.add_argument("--allow-dup-rename", action="store_true",
                           dest="allow_dup_rename")
    p_at_xlsx.add_argument("name_pos", nargs="?", default=None)
    p_at_xlsx.add_argument("headers_pos", nargs="?", default=None)
    p_at_xlsx.add_argument("rows_pos", nargs="*", default=[])
    p_at_xlsx.set_defaults(func=cmd_add_sheet)

    p_aa = sp.add_parser(
        "add-assumptions",
        help="Append a 3-col Item/Value/Source sheet for model inputs")
    p_aa.add_argument("file")
    p_aa.add_argument("--name", default=None,
                      help="sheet name (default 'Assumptions')")
    p_aa.add_argument("--title", default=None,
                      help="title row (default 'Assumptions')")
    p_aa.add_argument("--kv", action="append", default=[],
                      help="repeatable 'Item;Value;Source' "
                           f"(<= {_LIM['kv_max']} items)")
    p_aa.add_argument("--numfmt", action="append", default=[],
                      help="repeatable 'COL:FMT' (typically B:0.00% etc.)")
    p_aa.add_argument("--header-fill", default=None,
                      help="hex fill for header band (default 1F3864)")
    p_aa.set_defaults(func=cmd_add_assumptions)

    p_ad = sp.add_parser(
        "add-dashboard",
        help="Append a 2-col Metric/Value KPI sheet (formulas welcome)")
    p_ad.add_argument("file")
    p_ad.add_argument("--name", default=None,
                      help="sheet name (default 'Dashboard')")
    p_ad.add_argument("--title", default=None,
                      help="title row (default 'Dashboard')")
    p_ad.add_argument("--kpi", action="append", default=[],
                      help="repeatable 'Value_or_Formula;Label' "
                           f"(<= {_LIM['kpi_max']} items). Value first to "
                           "avoid shell-quoting traps around '='.")
    p_ad.add_argument("--header-fill", default=None,
                      help="hex fill for header band (default 1F3864)")
    p_ad.set_defaults(func=cmd_add_dashboard)

    # -- legacy / single-cell tools --------------------------------------
    p_ins = sp.add_parser("inspect", help="Per-sheet structure preview")
    p_ins.add_argument("file")
    p_ins.add_argument("--rows", type=int, default=8,
                       help="rows to preview per sheet (default 8)")
    p_ins.add_argument("--sheet", default=None, help="limit to one sheet")
    p_ins.set_defaults(func=cmd_inspect)

    p_ext = sp.add_parser("extract", help="CSV-like text dump")
    p_ext.add_argument("file")
    p_ext.add_argument("--sheet", default=None)
    p_ext.set_defaults(func=cmd_extract)

    # set-cell FILE SHEET A1 VALUE  (canonical, positional)
    p_st = sp.add_parser("set-cell",
                         help="Set a single cell: FILE SHEET A1 VALUE")
    p_st.add_argument("file")
    # Positional form is canonical; legacy --sheet/--cell still accepted
    # (they feed the same attrs, resolved by a small compat shim below).
    p_st.add_argument("sheet_pos", nargs="?", default=None,
                      help="sheet name")
    p_st.add_argument("cell_pos", nargs="?", default=None,
                      help="cell address e.g. 'A1', 'B12'")
    p_st.add_argument("value_pos", nargs="?", default=None,
                      help="value; '=' prefix -> formula, number literal "
                           "-> number, else text")
    p_st.add_argument("--numfmt", default=None,
                      help="number format, e.g. '#,##0' '0.0%' '$#,##0.00'")
    # Legacy flags (hidden from help; still functional).
    p_st.add_argument("--sheet", default=None, help=argparse.SUPPRESS)
    p_st.add_argument("--cell", default=None, help=argparse.SUPPRESS)
    p_st.add_argument("--text", default=None, help=argparse.SUPPRESS)
    p_st.add_argument("--number", default=None, help=argparse.SUPPRESS)
    p_st.add_argument("--formula", default=None, help=argparse.SUPPRESS)
    p_st.add_argument("--out", "--output", dest="out", default=None,
                      help=argparse.SUPPRESS)
    p_st.set_defaults(func=cmd_set_cell)

    # set-range FILE SHEET A1 "v1,v2,v3" ["v4,v5,v6" ...]
    # (canonical, positional; one CSV string per ROW, same as add-sheet/add-row)
    p_sr = sp.add_parser("set-range",
                         help=("Bulk write a rectangular area: "
                               "FILE SHEET A1 \"v1,v2,...\" [\"v3,v4,...\" ...]"))
    p_sr.add_argument("file")
    p_sr.add_argument("sheet_pos", nargs="?", default=None,
                      help="sheet name")
    p_sr.add_argument("start_pos", nargs="?", default=None,
                      help="anchor cell e.g. 'A1'")
    p_sr.add_argument("rows_pos", nargs="*", default=[],
                      help=("one CSV string per row; '=' prefix = formula, "
                            "numbers auto-typed, empty cell preserves existing"))
    # Legacy flags (hidden).
    p_sr.add_argument("--sheet", default=None, help=argparse.SUPPRESS)
    p_sr.add_argument("--start", default=None, help=argparse.SUPPRESS)
    p_sr.add_argument("--out", "--output", dest="out", default=None,
                      help=argparse.SUPPRESS)
    p_sr.set_defaults(func=cmd_set_range)

    p_nw = sp.add_parser("new-from-outline",
                         help="Build a multi-sheet xlsx from outline JSON")
    p_nw.add_argument("outline_pos", nargs="?", default=None,
                      help="outline JSON path (positional)")
    p_nw.add_argument("--outline", "--input", dest="outline_opt", default=None,
                      help="alias for the positional outline path")
    p_nw.add_argument("--out", "--output", dest="out", default=None,
                      help="output .xlsx (omit when --dry-run)")
    p_nw.add_argument("--auto-color", action="store_true",
                      help="apply financial color standard "
                           "(blue=hardcoded, black=formula)")
    p_nw.add_argument("--dry-run", action="store_true",
                      help="validate outline only; no file written")
    p_nw.set_defaults(func=cmd_new_from_outline)

    p_ex = sp.add_parser("examples",
                         help="List or dump built-in outline JSON examples")
    p_ex.add_argument("name", nargs="?", default="list",
                      help=f"one of: {list(EXAMPLES_XLSX)} or 'list'")
    p_ex.add_argument("--out", "--output", dest="out", default=None,
                      help="write example outline JSON to this file "
                           "(default: print to stdout)")
    p_ex.set_defaults(func=cmd_examples)

    p_cr = sp.add_parser("create",
                         help="Quick path: generate a workbook from a template in one shot")
    p_cr.add_argument("--template", required=True,
                      help=f"one of: {list(EXAMPLES_XLSX)}")
    p_cr.add_argument("--out", "--output", dest="out", required=True,
                      help="output .xlsx path")
    p_cr.add_argument("--outline", "--input", dest="outline_file",
                      default=None,
                      help="optional outline JSON that overrides the template")
    p_cr.add_argument("--auto-color", action="store_true",
                      help="apply financial color standard")
    p_cr.set_defaults(func=cmd_create)

    # [C] parse_known_args instead of parse_args: tolerate stray/typo flags
    # (e.g. --preset on add-sheet, misplaced --auto-color) with a visible
    # [WARN] instead of failing the whole command. argparse's strict mode
    # turned simple typos into exit-2 failures; the model then retried with
    # slightly different typos, producing the 9 Python FAILED lines in
    # log2.txt (Step 19/22/23/27/29/31/34/37). The warning line here keeps
    # the signal visible so the model still knows to correct the form.
    args, unknown = ap.parse_known_args()
    if unknown:
        sys.stderr.write(
            f"[WARN] ignoring unknown args: {unknown}. "
            f"These did not match any flag of the '{args.cmd}' subcommand.\n"
        )
    try:
        args.func(args)
        # Print the commands menu so the model can pick the next call.
        # We do NOT presume to suggest a specific next step -- the helper
        # has no idea what the overall task is.
        _print_commands_menu()
    except Exception as e:
        sys.stderr.write(f"[ERROR] {type(e).__name__}: {e}\n")
        _print_commands_menu(stream=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
